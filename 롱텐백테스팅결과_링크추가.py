import openpyxl
from openpyxl.styles import Font, Alignment


def create_excel_hyperlinks_all_sheets():
    file_name = "롱텐_종합백테스팅_최종결과.xlsx"
    output_name = "롱텐_종합백테스팅_링크추가_완성본.xlsx"

    try:
        wb = openpyxl.load_workbook(file_name)
        ws_log = wb["상세로그"]

        # 엑셀 열 때 자동 계산 활성화
        wb.calculation.calcMode = "auto"

        # ==========================================
        # 1. 상세로그 시트 사전 탐색 (매핑 딕셔너리 생성)
        # ==========================================
        log_map = {}
        for r in range(2, ws_log.max_row + 1):
            name = ws_log.cell(row=r, column=1).value
            date_val = ws_log.cell(row=r, column=2).value

            name_str = str(name).strip() if name else ""

            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val).strip().split(" ")[0]

            key = f"{name_str}_{date_str}"
            if name_str and key not in log_map:
                log_map[key] = r

        # [가독성] 상세로그 시트 셀 크기(열 너비) 조정
        log_widths = {'A': 16, 'B': 14, 'C': 15, 'D': 12, 'E': 65}
        for col_letter, width in log_widths.items():
            ws_log.column_dimensions[col_letter].width = width

        # ==========================================
        # 2. 3개의 요약 시트에 순차적으로 하이퍼링크 삽입
        # ==========================================
        target_sheets = ["매매요약(종목순)", "매수일별정렬", "미청산(보유중)"]

        # 요약 시트용 최적화된 열 너비
        summary_widths = {
            'A': 16, 'B': 10, 'C': 14, 'D': 10, 'E': 12,
            'F': 12, 'G': 12, 'H': 12, 'I': 10, 'J': 10,
            'K': 10, 'L': 10, 'M': 50
        }

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                continue  # 만약 파일에 해당 시트가 없으면 건너뜀

            ws = wb[sheet_name]

            # C열(3번째 열)에 '상세로그링크' 열 삽입
            ws.insert_cols(3)
            ws.cell(row=1, column=3, value="상세로그링크")

            for r in range(2, ws.max_row + 1):
                name = ws.cell(row=r, column=1).value
                # C열 삽입으로 인해 1차매수일(D열)이 E열(5번째)로 이동함
                date_val = ws.cell(row=r, column=5).value

                name_str = str(name).strip() if name else ""

                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val).strip().split(" ")[0]

                key = f"{name_str}_{date_str}"
                target_row = log_map.get(key)

                # 링크 및 스타일 삽입
                if target_row:
                    safe_formula = f'=HYPERLINK("#\'상세로그\'!A{target_row}", "🔍 이동")'
                    ws.cell(row=r, column=3, value=safe_formula)
                    ws.cell(row=r, column=3).font = Font(color="0000FF", underline="single")
                else:
                    ws.cell(row=r, column=3, value="-")  # 매칭 안 되면 하이픈 표시

                # 링크 열(C열) 글자 가운데 정렬
                ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')

            # [가독성] 요약 시트 셀 크기(열 너비) 조정
            for col_letter, width in summary_widths.items():
                ws.column_dimensions[col_letter].width = width

        # ==========================================
        # 3. 완성본 저장
        # ==========================================
        wb.save(output_name)
        print(f"🎉 성공! 3개 시트에 링크가 연결되고 가독성이 최적화되었습니다: {output_name}")

    except PermissionError:
        print("\n❌ 파일 열림 오류: 엑셀 파일을 완전히 닫고 코드를 다시 실행해주세요.\n")
    except Exception as e:
        print(f"\n❌ 기타 에러 발생: {e}\n")


if __name__ == "__main__":
    create_excel_hyperlinks_all_sheets()