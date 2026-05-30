import os
import sqlite3
import pandas as pd
from calendar import monthrange
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ==========================================
# [설정 항목] 파일 경로 고정 정의
# ==========================================
CSV_CANDIDATE_PATH = "까리존345후보종목.csv"
DB_PATH = "stock_data.db"
EXCEL_OUTPUT_PATH = "까리존345_장중터치_보수적결과.xlsx"

UNIT_MONEY = 1000000  # 1차 진입 원금: 100만 원
END_SIMULATION_DATE = "2026-05-29"  # 최종 추적 종료일


def get_user_date_range():
    print("=" * 80)
    print(" [까리존 345 - 장중 터치 기반 보수적 동시터치 방어 엔진 V2 (수수료 제외)]")
    print("=" * 80)
    while True:
        start_in = input("▶ 테스트 시작 연월을 입력하세요 (예: 202001): ").replace("-", "").strip()
        if len(start_in) == 6 and start_in.isdigit():
            start_date = f"{start_in[:4]}-{start_in[4:6]}-01"
            break
        print("[입력오류] YYYYMM 형식으로 정확히 입력해 주세요.")
    while True:
        end_in = input("▶ 테스트 종료 연월을 입력하세요 (예: 202512): ").replace("-", "").strip()
        if len(end_in) == 6 and end_in.isdigit():
            year = int(end_in[:4])
            month = int(end_in[4:6])
            if 1 <= month <= 12:
                last_day = monthrange(year, month)[1]
                end_date = f"{end_in[:4]}-{end_in[4:6]}-{last_day}"
                break
        print("[입력오류] YYYYMM 형식으로 정확히 입력해 주세요.")
    return start_date, end_date


def load_target_candidates(start_date, end_date):
    if not os.path.exists(CSV_CANDIDATE_PATH):
        raise FileNotFoundError(f"후보 종목 파일이 없습니다: {CSV_CANDIDATE_PATH}")
    df = pd.read_csv(CSV_CANDIDATE_PATH, dtype={"code": str, "date": str})
    if df.empty: return df
    sample_date = str(df["date"].iloc[0])
    cmp_start = start_date.replace("-", "") if "-" not in sample_date else start_date
    cmp_end = end_date.replace("-", "") if "-" not in sample_date else end_date
    return df[(df["date"] >= cmp_start) & (df["date"] <= cmp_end)].copy()


def detect_db_date_format(conn):
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT date FROM daily_prices LIMIT 1")
        row = cursor.fetchone()
        if row and row[0]: return "HYPHEN" if "-" in str(row[0]) else "RAW"
    except:
        pass
    return "HYPHEN"


def main():
    start_candidate_date, end_candidate_date = get_user_date_range()
    print(f"\n[시스템] 엔진 빌드를 시작합니다 (장중 터치 즉시 체결 및 최악 가설 동시터치 룰 적용)...")
    try:
        df_targets = load_target_candidates(start_candidate_date, end_candidate_date)
        total_event_count = len(df_targets)
        print(f"[안내] 후보 이벤트 개수: 총 {total_event_count}건 로드 완료")
        if df_targets.empty: return
    except Exception as e:
        print(f"[오류] 후보 로드 실패: {e}")
        return

    conn = sqlite3.connect(DB_PATH)
    db_date_type = detect_db_date_format(conn)

    cursor = conn.cursor()
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_daily_prices_code_date_high ON daily_prices (code, date DESC, high);")
    conn.commit()

    def normalize_to_db_format(x):
        pure_date = str(x).replace("-", "").strip()
        return f"{pure_date[:4]}-{pure_date[4:6]}-{pure_date[6:8]}" if db_date_type == "HYPHEN" else pure_date

    df_targets["date"] = df_targets["date"].apply(normalize_to_db_format)
    db_start_date = normalize_to_db_format(start_candidate_date)
    db_end_date = normalize_to_db_format(END_SIMULATION_DATE)

    cursor.execute("SELECT DISTINCT date FROM daily_prices WHERE date >= ? AND date <= ? ORDER BY date ASC",
                   (db_start_date, db_end_date))
    timeline = [r[0] for r in cursor.fetchall()]

    portfolio = {}
    pending_signals = {}
    all_target_audit_logs = {}

    df_targets = df_targets.sort_values(by="date").reset_index(drop=True)
    for _, row in df_targets.iterrows():
        code_str = str(row["code"]).zfill(6)
        target_date = row["date"]

        cursor.execute("SELECT high FROM daily_prices WHERE code = ? AND date = ?", (code_str, target_date))
        res = cursor.fetchone()
        if res:
            base_high = res[0]
            p_33_price = int(base_high * 0.67)
            cursor.execute("SELECT COUNT(*) FROM daily_prices WHERE code = ? AND date <= ?", (code_str, target_date))
            history_count = cursor.fetchone()[0]
            if history_count < 240:
                all_target_audit_logs[(code_str, target_date)] = {
                    "후보등록일": target_date, "종목명": row["name"], "종목코드": code_str, "기준봉고가": base_high,
                    "검증일기준고점": base_high, "타점가격(-33%)": p_33_price,
                    "최종상태": "매수제외(상장기간부족)", "최초매수일": "-", "최종매도일": "-", "총투자금액": 0, "수익금": 0
                }
                continue

            pending_signals[(code_str, target_date)] = {
                "name": row["name"], "base_high": base_high, "active": True, "wait_days": 0
            }

            base_audit = {
                "후보등록일": target_date, "종목명": row["name"], "종목코드": code_str, "기준봉고가": base_high,
                "검증일기준고점": base_high, "타점가격(-33%)": p_33_price,
                "최종상태": "미체결(감시중)", "최초매수일": "-", "최종매도일": "-", "총투자금액": 0, "수익금": 0
            }
            all_target_audit_logs[(code_str, target_date)] = base_audit

    peak_capital = 0
    trade_logs = []

    for current_date in timeline:
        cleared_codes = []

        # =========================================================================
        # [A. 보유 종목 자동 청산 & 추매 로직 (장중 터치 우선순위 엔진)]
        # =========================================================================
        for code, stock in portfolio.items():
            cursor.execute("SELECT open, high, low, close FROM daily_prices WHERE code = ? AND date = ?",
                           (code, current_date))
            day_data = cursor.fetchone()
            if not day_data: continue
            s_open, s_high, s_low, s_close = day_data

            stock["holding_days"] += 1
            signal_date = stock["signal_date"]

            cursor.execute(
                "SELECT MAX(high) FROM (SELECT high FROM daily_prices WHERE code = ? AND date <= ? ORDER BY date DESC LIMIT 120)",
                (code, current_date))
            rolling_res = cursor.fetchone()
            current_rolling_high = rolling_res[0] if (rolling_res and rolling_res[0]) else stock["base_high"]

            # 추매 타점 계산
            p_43_price = current_rolling_high * 0.57
            p_50_price = current_rolling_high * 0.50
            p_66_price = current_rolling_high * 0.34

            old_avg_price = stock["avg_price"]

            # [기본 청산 목표가 설정 (수수료 제외 버전)]
            if len(stock["triggered_phases"]) > 1 or stock["status"] == "WAIT":
                target_exit_price = old_avg_price
                exit_type_base = "본절마감(일반)"
                profit = 0
            else:
                target_exit_price = old_avg_price * 1.06
                exit_type_base = "익절청산(+6%)"
                profit = int(stock["total_spent"] * 0.06)

            # -------------------------------------------------------------
            # 1. 시가 갭돌파 판별 (최우선 즉시 청산)
            # -------------------------------------------------------------
            if s_open >= target_exit_price:
                trade_logs.append({
                    "종목명": stock["name"], "종목코드": code, "청산유형": f"{exit_type_base}(시가갭)",
                    "최종투자금": stock["total_spent"], "수익금": profit, "보유일수": stock["holding_days"],
                    "매수일자": stock["buy_date"], "매도일자": current_date, "추매경로기록": ", ".join(stock["history_logs"])
                })
                all_target_audit_logs[(code, signal_date)].update({
                    "최종상태": f"매매완료({exit_type_base.split('(')[0]})", "최초매수일": stock["buy_date"], "최종매도일": current_date,
                    "총투자금액": stock["total_spent"], "수익금": profit
                })
                cleared_codes.append(code)
                continue

            # -------------------------------------------------------------
            # 2. 장중 하방 터치 판별 (추매 집행 유무 확인 - 최악 가설 선반영)
            # -------------------------------------------------------------
            pyramided_today = False

            if "43" not in stock["triggered_phases"] and s_low <= p_43_price:
                add_spent = int(UNIT_MONEY * 1.1)
                stock["total_spent"] += add_spent
                stock["total_qty"] += (add_spent / p_43_price)  # 지정가 터치 체결 처리
                stock["triggered_phases"].append("43")
                stock["history_logs"].append(f"[-43%추매(기준고점:{int(current_rolling_high)})]:{current_date}")
                pyramided_today = True

            if "50" not in stock["triggered_phases"] and s_low <= p_50_price:
                add_spent = int(UNIT_MONEY * 4.3)
                stock["total_spent"] += add_spent
                stock["total_qty"] += (add_spent / p_50_price)
                stock["triggered_phases"].append("50")
                stock["history_logs"].append(f"[-50%추매(기준고점:{int(current_rolling_high)})]:{current_date}")
                pyramided_today = True

            if "66" not in stock["triggered_phases"] and s_low <= p_66_price:
                add_spent = int(UNIT_MONEY * 7.0)
                stock["total_spent"] += add_spent
                stock["total_qty"] += (add_spent / p_66_price)
                stock["triggered_phases"].append("66")
                stock["history_logs"].append(f"[-66%추매(기준고점:{int(current_rolling_high)})]:{current_date}")
                pyramided_today = True

            # 당일 추매가 발생했다면 평단가를 낮추고, 청산 목표가(본절가) 즉시 수정
            if pyramided_today:
                stock["avg_price"] = stock["total_spent"] / stock["total_qty"]
                target_exit_price = stock["avg_price"]
                exit_type_base = "본절마감(장중동시터치)"
                profit = 0

            # -------------------------------------------------------------
            # 3. 장중 상방 터치 판별 (변경된 목표가 기준으로 최종 청산 감시)
            # -------------------------------------------------------------
            if s_high >= target_exit_price:
                trade_logs.append({
                    "종목명": stock["name"], "종목코드": code, "청산유형": exit_type_base,
                    "최종투자금": stock["total_spent"], "수익금": profit, "보유일수": stock["holding_days"],
                    "매수일자": stock["buy_date"], "매도일자": current_date, "추매경로기록": ", ".join(stock["history_logs"])
                })
                all_target_audit_logs[(code, signal_date)].update({
                    "최종상태": f"매매완료({exit_type_base.split('(')[0]})", "최초매수일": stock["buy_date"], "최종매도일": current_date,
                    "총투자금액": stock["total_spent"], "수익금": profit
                })
                cleared_codes.append(code)
                continue

            # 당일 청산이 안 되고 10일 차가 넘어가면 WAIT(본절 대기) 상태로 전환
            if stock["holding_days"] == 10 and stock["status"] == "HOLD" and code not in cleared_codes:
                stock["status"] = "WAIT"

            all_target_audit_logs[(code, signal_date)].update({
                "최종상태": f"보유중({stock['status']})", "최초매수일": stock["buy_date"], "총투자금액": stock["total_spent"]
            })

        for c_code in cleared_codes: del portfolio[c_code]

        # =========================================================================
        # [B-1. 신형 기준봉 출현 시 기존 신호 교체 무효화 로직]
        # =========================================================================
        for (code, s_date), signal in list(pending_signals.items()):
            if current_date != s_date: continue
            if not signal["active"]: continue

            for (old_code, old_date), old_signal in pending_signals.items():
                if old_code == code and old_date < s_date and old_signal["active"]:
                    if code not in portfolio and old_signal["wait_days"] <= 10:
                        old_signal["active"] = False
                        all_target_audit_logs[(old_code, old_date)].update({
                            "최종상태": f"미체결(신형기준봉 교체 폐기, 신형일자:{s_date})"
                        })

        # =========================================================================
        # [B-2. 신규 종목 장중 터치 1차 진입 감시]
        # =========================================================================
        for (code, s_date), signal in pending_signals.items():
            if not signal["active"]: continue
            if current_date <= s_date: continue

            signal["wait_days"] += 1
            if signal["wait_days"] > 120:
                signal["active"] = False
                if all_target_audit_logs[(code, s_date)]["최종상태"] == "미체결(감시중)":
                    all_target_audit_logs[(code, s_date)].update({"최종상태": "미체결(6개월기간만료)"})
                continue

            if code in portfolio: continue

            cursor.execute(
                "SELECT MAX(high) FROM (SELECT high FROM daily_prices WHERE code = ? AND date <= ? ORDER BY date DESC LIMIT 120)",
                (code, current_date))
            rolling_res = cursor.fetchone()
            rolling_high = rolling_res[0] if (rolling_res and rolling_res[0]) else signal["base_high"]

            cursor.execute("SELECT open, high, low, close FROM daily_prices WHERE code = ? AND date = ?",
                           (code, current_date))
            d_data = cursor.fetchone()
            if not d_data: continue
            d_open, d_high, d_low, d_close = d_data

            p_33_price = rolling_high * 0.67
            p_50_price = rolling_high * 0.50

            all_target_audit_logs[(code, s_date)]["검증일기준고점"] = int(rolling_high)
            all_target_audit_logs[(code, s_date)]["타점가격(-33%)"] = int(p_33_price)

            # 진입 전 관삭: 시가 갭하락으로 -50%를 뚫고 시작하거나 장중 저가가 -50% 이하까지 밀린 경우
            if d_open <= p_50_price or d_low <= p_50_price:
                signal["active"] = False
                all_target_audit_logs[(code, s_date)].update({"최종상태": "미체결(진입전관삭)"})
                continue

            # 장중 -33% 지정가 터치 매수
            if d_low <= p_33_price:
                signal["active"] = False
                phases = ["33"]
                total_spent = UNIT_MONEY
                history_list = [f"[-33%장중터치진입(기준고점:{int(rolling_high)})]:{current_date}"]

                portfolio[code] = {
                    "name": signal["name"], "base_high": rolling_high, "holding_days": 1,
                    "total_spent": total_spent, "total_qty": total_spent / p_33_price, "avg_price": p_33_price,
                    "status": "HOLD", "triggered_phases": phases, "buy_date": current_date,
                    "history_logs": history_list,
                    "signal_date": s_date
                }
                all_target_audit_logs[(code, s_date)].update({
                    "최종상태": "보유중(HOLD)", "최초매수일": current_date, "총투자금액": total_spent
                })

        current_total_invested = sum(s["total_spent"] for s in portfolio.values())
        if current_total_invested > peak_capital: peak_capital = current_total_invested

    conn.close()

    # ==========================================
    # [데이터 프레임 생성 및 엑셀 출력 (가독성 최적화)]
    # ==========================================
    df_summary = pd.DataFrame({
        "지표명": [f"선택 기간 총 후보 건수", "매매 완료 청산 건수", "현재 미청산 보유 건수", "역대 최고 자금 요구치 (Peak Capital)"],
        "수치 데이터": [f"{total_event_count} 건", f"{len(trade_logs)} 건", f"{len(portfolio)} 건", f"{peak_capital:,} 원"]
    })
    df_trade_logs = pd.DataFrame(trade_logs) if trade_logs else pd.DataFrame(
        columns=["종목명", "종목코드", "청산유형", "최종투자금", "수익금", "보유일수", "매수일자", "매도일자", "추매경로기록"])
    df_audit_sheet = pd.DataFrame(all_target_audit_logs.values())
    if not df_audit_sheet.empty:
        df_audit_sheet = df_audit_sheet.sort_values(by=["후보등록일", "종목명"]).reset_index(drop=True)

    with pd.ExcelWriter(EXCEL_OUTPUT_PATH, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="종합요약통계", index=False)
        df_trade_logs.to_excel(writer, sheet_name="매매완료상세내역", index=False)
        df_audit_sheet.to_excel(writer, sheet_name="전수조사종목현황", index=False)

        workbook = writer.book
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        font_header = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="맑은 고딕", size=10, bold=False)
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            worksheet.views.sheetView[0].showGridLines = True

            for col_idx, col in enumerate(worksheet.columns, start=1):
                max_len = 0
                col_letter = get_column_letter(col_idx)

                for cell in col:
                    cell.font = font_body
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 100000:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = align_center
                    else:
                        if sheet_name == "매매완료상세내역" and col_idx == 9:
                            cell.alignment = align_left
                        elif sheet_name == "전수조사종목현황" and col_idx >= 12:
                            cell.alignment = align_left
                        else:
                            cell.alignment = align_center

                    if cell.value is not None:
                        val_str = str(cell.value)
                        actual_len = sum(2 if ord(char) > 128 else 1 for char in val_str)
                        if actual_len > max_len: max_len = actual_len

                header_cell = worksheet[f"{col_letter}1"]
                header_cell.fill = fill_header
                header_cell.font = font_header
                header_cell.alignment = align_center
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    print("\n" + "=" * 115)
    print(f" [시스템] 장중 터치 기반 보수적 엔진 백테스트 완료 (수수료 미적용)")
    print(f" ▶ 최종 산출물 물리 저장 경로 : {EXCEL_OUTPUT_PATH}")
    print("=" * 115)


if __name__ == "__main__":
    main()