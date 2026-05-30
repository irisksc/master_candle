import sqlite3
import pandas as pd
import os
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


def run_advanced_switch_backtest():
    current_dir = os.path.dirname(os.path.abspath('%s' % __file__)) if '__file__' in locals() else os.getcwd()
    csv_path = os.path.join(current_dir, '롱텐관심종목(10년).csv')
    db_path = os.path.join(current_dir, '롱텐주봉데이터.db')
    excel_path = os.path.join(current_dir, '롱텐_종합백테스팅_최종결과.xlsx')

    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"'{csv_path}' 파일이 없습니다.")

    signal_df = pd.read_csv(csv_path)
    signal_df['날짜'] = pd.to_datetime(signal_df['날짜'])
    unique_codes = signal_df['종목코드'].unique()

    print(f"\n" + "=" * 75)
    print(f"▶ 롱텐 엑셀 마스터 리포트 (동적 Envelope 추매 무손절 버전) 생성 중...")
    print("=" * 75)

    summary_data = []
    detail_data = []

    stat_total_trades = 0
    stat_clean_exits = 0
    stat_profit_exits = 0

    conn = sqlite3.connect(db_path)

    for code in tqdm(unique_codes, desc="백테스팅 진행 및 데이터 수집"):
        code_str = str(code).zfill(6)
        stock_name = signal_df[signal_df['종목코드'] == code]['종목명'].iloc[0]
        stock_signals = signal_df[signal_df['종목코드'] == code]['날짜'].values

        df_prices = pd.read_sql(
            f"SELECT weekly_start_date, open, high, low, close FROM weekly_prices WHERE code='{code_str}' ORDER BY weekly_start_date ASC",
            conn
        )

        if df_prices.empty: continue

        df_prices['weekly_start_date'] = pd.to_datetime(df_prices['weekly_start_date'])
        df_prices.set_index('weekly_start_date', inplace=True)

        # 10주 이동평균선 및 동적 Envelope 밴드 계산
        df_prices['MA10'] = df_prices['close'].rolling(window=10).mean()
        df_prices['env10'] = df_prices['MA10'] * 0.90
        df_prices['env20'] = df_prices['MA10'] * 0.80
        df_prices['env30'] = df_prices['MA10'] * 0.70
        df_prices['env40'] = df_prices['MA10'] * 0.60

        is_holding = False
        base_price, total_cash, total_qty, buy_stage = 0, 0, 0, 0
        entry_date = None
        last_buy_band = None  # 경로 의존형 밴드 추적 변수

        trade_peak = 0
        trade_mdd = 0
        buy_history_list = []

        for current_date, price_row in df_prices.iterrows():
            date_str = current_date.strftime('%Y-%m-%d')

            ma10 = price_row['MA10']

            # [해결책 1 반영] 지표 생성 전(최초 9주)의 NaN 구간에서는 매매 판단 패스
            if pd.isna(ma10):
                continue

            # ==========================================
            # [A] 미보유 상태: 시그널에 따른 최초 진입 (0차)
            # ==========================================
            if not is_holding:
                if current_date in stock_signals:
                    stat_total_trades += 1
                    base_price = price_row['close']  # 종가 진입
                    total_cash = 10_000_000
                    total_qty = total_cash / base_price
                    buy_stage = 0
                    last_buy_band = None
                    is_holding = True
                    entry_date = current_date

                    trade_peak = base_price
                    trade_mdd = 0
                    buy_history_list = []

                    detail_data.append({
                        '종목명': stock_name, '날짜': date_str, '액션': '★ 0차 진입',
                        '평단가': int(base_price), '상세내용': f"시그널 진입가: {int(base_price):,}원"
                    })
                continue

            # ==========================================
            # [B] 보유 상태: 청산 및 동적 추매 관리
            # ==========================================
            if is_holding:
                high_p = price_row['high']
                low_p = price_row['low']
                curr_close = price_row['close']

                env10 = price_row['env10']
                env20 = price_row['env20']
                env30 = price_row['env30']
                env40 = price_row['env40']

                avg_price = total_cash / total_qty

                # 통계용 MDD 갱신
                if high_p > trade_peak: trade_peak = high_p
                current_dd = (low_p - trade_peak) / trade_peak * 100
                if current_dd < trade_mdd: trade_mdd = current_dd

                duration_days = (current_date - entry_date).days

                # ------------------------------------------
                # 1. 장중 본절 탈출 감시 (추매 1회 이상 시 활성화)
                # ------------------------------------------
                if buy_stage >= 1 and high_p >= avg_price:
                    stat_clean_exits += 1
                    summary_data.append({
                        '종목명': stock_name, '종목번호': code_str, '결과': '본절탈출',
                        '1차매수일': entry_date.strftime('%Y-%m-%d'), '매수가': int(base_price),
                        '매도일': date_str, '매도가': int(avg_price),
                        '수익률(%)': 0.0, '기간(일)': duration_days, 'MDD(%)': round(trade_mdd, 2),
                        '추매횟수': f"{buy_stage}차", '추매내용': " -> ".join(buy_history_list)
                    })
                    detail_data.append({
                        '종목명': stock_name, '날짜': date_str, '액션': '♣ 본절청산',
                        '평단가': int(avg_price), '상세내용': f"장중 고가({int(high_p):,}원)가 낮아진 평단가({int(avg_price):,}원) 터치 (탈출)"
                    })
                    is_holding = False
                    continue

                # ------------------------------------------
                # 2. 금요일 종가 감시 (0차: 익절 or 1차 추매)
                # ------------------------------------------
                if buy_stage == 0:
                    if curr_close < env10:
                        if curr_close >= base_price:
                            # [조건 B] 엔벨로프 익절
                            stat_profit_exits += 1
                            roi = ((curr_close - base_price) / base_price) * 100
                            summary_data.append({
                                '종목명': stock_name, '종목번호': code_str, '결과': '엔벨익절',
                                '1차매수일': entry_date.strftime('%Y-%m-%d'), '매수가': int(base_price),
                                '매도일': date_str, '매도가': int(curr_close),
                                '수익률(%)': round(roi, 2), '기간(일)': duration_days, 'MDD(%)': round(trade_mdd, 2),
                                '추매횟수': "0차", '추매내용': "추매 없음(수익 확정)"
                            })
                            detail_data.append({
                                '종목명': stock_name, '날짜': date_str, '액션': '엔벨익절',
                                '평단가': int(base_price), '상세내용': f"Env 10 이탈이나 진입가보다 높아 수익 청산"
                            })
                            is_holding = False
                            continue
                        else:
                            # 익절 Pass 및 1차 추매 동시 집행 (손실 구간에서 Env 10 이탈)
                            exec_price = curr_close
                            added_cash = 10_000_000 * 1.5
                            total_qty += added_cash / exec_price
                            total_cash += added_cash
                            buy_stage = 1
                            avg_price = total_cash / total_qty

                            # 차기 타깃을 위한 현재 밴드 파악
                            if curr_close < env40:
                                last_buy_band = 'env40'
                            elif curr_close < env30:
                                last_buy_band = 'env30'
                            elif curr_close < env20:
                                last_buy_band = 'env20'
                            else:
                                last_buy_band = 'env10'

                            date_short = f"{current_date.month}/{current_date.day}"
                            buy_history_list.append(f"1차({date_short}) {int(exec_price)} 평단 {int(avg_price)}")
                            detail_data.append({
                                '종목명': stock_name, '날짜': date_str, '액션': '└─ 1차 추매',
                                '평단가': int(avg_price), '상세내용': f"손실 중 Env 10 이탈 (종가 {int(exec_price):,}원) 1.5배 체결"
                            })

                # ------------------------------------------
                # 3. 금요일 종가 감시 (1차 이상: 경로 의존형 동적 추매)
                # ------------------------------------------
                elif buy_stage in [1, 2, 3]:
                    target_map = {'env10': env20, 'env20': env30, 'env30': env40, 'env40': None}
                    target_price = target_map.get(last_buy_band)

                    if target_price and curr_close < target_price:
                        exec_price = curr_close

                        # 차수에 따른 엔젤 배수 적용
                        if buy_stage == 1:
                            added_cash = 10_000_000 * 2.0
                        elif buy_stage == 2:
                            added_cash = 10_000_000 * 2.5
                        elif buy_stage == 3:
                            added_cash = 10_000_000 * 3.0

                        total_qty += added_cash / exec_price
                        total_cash += added_cash
                        buy_stage += 1
                        avg_price = total_cash / total_qty

                        # 체결 후 차기 타깃 갱신
                        prev_band = last_buy_band
                        if curr_close < env40:
                            last_buy_band = 'env40'
                        elif curr_close < env30:
                            last_buy_band = 'env30'
                        elif curr_close < env20:
                            last_buy_band = 'env20'
                        else:
                            last_buy_band = 'env10'

                        date_short = f"{current_date.month}/{current_date.day}"
                        buy_history_list.append(f"{buy_stage}차({date_short}) {int(exec_price)} 평단 {int(avg_price)}")
                        detail_data.append({
                            '종목명': stock_name, '날짜': date_str, '액션': f'└─ {buy_stage}차 추매',
                            '평단가': int(avg_price), '상세내용': f"직전 밴드({prev_band}) 하향 갱신 (종가 {int(exec_price):,}원) 체결"
                        })

        # ==========================================
        # [C] 미청산(보유중) 기록 처리
        # ==========================================
        if is_holding:
            avg_price = total_cash / total_qty
            roi = ((df_prices['close'].iloc[-1] - avg_price) / avg_price) * 100
            duration_days = (df_prices.index[-1] - entry_date).days

            summary_data.append({
                '종목명': stock_name, '종목번호': code_str, '결과': '보유중',
                '1차매수일': entry_date.strftime('%Y-%m-%d'), '매수가': int(base_price),
                '매도일': '-', '매도가': int(df_prices['close'].iloc[-1]),
                '수익률(%)': round(roi, 2), '기간(일)': duration_days, 'MDD(%)': round(trade_mdd, 2),
                '추매횟수': f"{buy_stage}차", '추매내용': " -> ".join(buy_history_list) if buy_history_list else "추매 대기"
            })

    conn.close()

    # ---------------------------------------------------------
    # 데이터프레임 변환 및 엑셀 포매팅 후처리
    # ---------------------------------------------------------
    df_summary = pd.DataFrame(summary_data)
    df_detail = pd.DataFrame(detail_data)

    df_summary_sorted = df_summary.copy()
    if not df_summary_sorted.empty:
        df_summary_sorted = df_summary_sorted.sort_values(by='1차매수일')

    df_holding = df_summary[df_summary['결과'] == '보유중'].copy()
    if not df_holding.empty:
        df_holding = df_holding.sort_values(by='수익률(%)', ascending=True)

    total_closed = stat_clean_exits + stat_profit_exits
    win_rate = (total_closed / stat_total_trades * 100) if stat_total_trades > 0 else 0

    stats_data = [{
        '총 매매 진입 횟수': stat_total_trades,
        '완료된 매매(청산)': total_closed,
        '미청산(보유중)': stat_total_trades - total_closed,
        '본절 탈출 횟수': stat_clean_exits,
        '엔벨로프 익절 횟수': stat_profit_exits,
        '종합 청산 승률 (%)': f"{round(win_rate, 2)}%"
    }]
    df_stats = pd.DataFrame(stats_data)

    fill_profit = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_breakeven = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    fill_holding = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    fill_alt1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_alt2 = PatternFill(start_color="F2F9FF", end_color="F2F9FF", fill_type="solid")

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_stats.to_excel(writer, sheet_name='통계요약', index=False)
        df_summary.to_excel(writer, sheet_name='매매요약(종목순)', index=False)
        df_summary_sorted.to_excel(writer, sheet_name='매수일별정렬', index=False)
        df_holding.to_excel(writer, sheet_name='미청산(보유중)', index=False)
        df_detail.to_excel(writer, sheet_name='상세로그', index=False)

        for sheet_name in ['매매요약(종목순)', '매수일별정렬', '미청산(보유중)']:
            ws = writer.sheets[sheet_name]
            res_col_idx = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == '결과':
                    res_col_idx = col
                    break

            if res_col_idx:
                for row in range(2, ws.max_row + 1):
                    val = ws.cell(row=row, column=res_col_idx).value
                    if val == '엔벨익절':
                        ws.cell(row=row, column=res_col_idx).fill = fill_profit
                    elif val == '본절탈출':
                        ws.cell(row=row, column=res_col_idx).fill = fill_breakeven
                    elif val == '보유중':
                        ws.cell(row=row, column=res_col_idx).fill = fill_holding

        ws_detail = writer.sheets['상세로그']
        action_col_idx = None
        for col in range(1, ws_detail.max_column + 1):
            if ws_detail.cell(row=1, column=col).value == '액션':
                action_col_idx = col
                break

        if action_col_idx:
            current_fill = fill_alt1
            for row in range(2, ws_detail.max_row + 1):
                action_val = str(ws_detail.cell(row=row, column=action_col_idx).value)
                if '★ 0차 진입' in action_val:
                    current_fill = fill_alt2 if current_fill == fill_alt1 else fill_alt1

                for col in range(1, ws_detail.max_column + 1):
                    ws_detail.cell(row=row, column=col).fill = current_fill

        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for col in worksheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            val_len = len(str(cell.value))
                            if any(ord(char) > 127 for char in str(cell.value)):
                                val_len = int(val_len * 1.5)
                            if val_len > max_length:
                                max_length = val_len
                    except:
                        pass
                worksheet.column_dimensions[col_letter].width = max_length + 2

    print("\n" + "=" * 75)
    print(f"■ 백테스팅 완료! 동적 추매 구조가 적용된 최종 엑셀 리포트가 생성되었습니다.")
    print("=" * 75)


if __name__ == "__main__":
    run_advanced_switch_backtest()